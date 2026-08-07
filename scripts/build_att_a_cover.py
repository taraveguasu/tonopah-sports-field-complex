#!/usr/bin/env python3
"""
Build the Attachment A Review Cover Sheet and the contract-amount summary that
sits behind it.

The Bluebeam process document defines the REVIEW packet: cover sheet first, then
"on a separate sheet, include a summary of how we got to the contract amount that
is on the Cover Sheet", then the draft Attachment A, then the bid form, the
subcontractor's own proposal, scope-option backup, and the descope notes.

PM direction: the ONLY place to break out pricing is this cover sheet. The
Attachment A itself carries one LUMP SUM.

The template is a legacy .xls, which openpyxl cannot write, so the cover sheet is
emitted as .xlsx with the same layout and labels. Noted rather than silently
swapped -- the file extension changes.

Usage:  python3 scripts/build_att_a_cover.py RFP-008
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

ROOT = Path(__file__).resolve().parent.parent
CONTENT = ROOT / "01-index" / "attachment-a-content"
OUT = ROOT / "02-drafts"

PROJECT = "NCSD - Tonopah High School Sports Field Replacement"
PROJECT_NO = "26-01-019"

BOLD = Font(bold=True)
THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def money(v):
    return f"${v:,.2f}"


def build(pkg):
    spec = json.loads((CONTENT / f"{pkg}.json").read_text())
    dest = OUT / pkg
    dest.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cover sheet"
    for col, w in (("A", 3), ("B", 34), ("C", 20), ("D", 22), ("E", 20), ("F", 18)):
        ws.column_dimensions[col].width = w

    ws["D2"] = "Attachment A Review Log"; ws["D2"].font = Font(bold=True, size=14)
    ws["D3"] = PROJECT
    ws["D4"] = PROJECT_NO
    for r in (3, 4):
        ws[f"D{r}"].font = BOLD

    ws["B6"], ws["C6"], ws["E6"], ws["F6"] = ("Approved By", "Approval Stamp",
                                              "Reviewed By", "Reviewed Stamp")
    for c in ("B6", "C6", "E6", "F6"):
        ws[c].font = BOLD
        ws[c].border = BOX
        ws[c].alignment = Alignment(horizontal="center")
    for i, (a, b) in enumerate([("Project Director", "Project Manager"),
                                ("General Superintendent", "Superintendent"),
                                ("Tim Roley or Matt Wade (if over $5M)",
                                 "Assistant PM (if applicable)")], start=7):
        ws[f"B{i}"], ws[f"E{i}"] = a, b
        for c in (f"B{i}", f"C{i}", f"E{i}", f"F{i}"):
            ws[c].border = BOX

    ws["B11"] = ("Final Session; add Invitees: Anne Tall, Erin Hicks, "
                 "Kathleen Hamilton, Liz Pippett, Inger Pippett")

    ws["B13"] = "Subcontractor:"; ws["B13"].font = BOLD
    ws["D13"] = spec["_subcontractor"]

    ws["B15"] = "Anticipated Material Procurement Date:"
    ws["D15"] = "CONFIRM"
    ws["B16"] = "Anticipated Start Date:"
    ws["D16"] = "CONFIRM"
    for c in ("D15", "D16"):
        ws[c].font = Font(bold=True, color="FF0000")

    ws["B19"] = "*Phase Code"
    ws["C19"] = spec.get("phase_code", "CONFIRM")
    ws["E19"] = "Dollar Amount"
    ws["F19"] = money(spec["_contract_amount"])
    ws["B20"] = "    *Only one Phase Code Per Subcontractor/Vendor per Tim Roley"

    ws["B27"] = "TOTAL (should equal contract total)"; ws["B27"].font = BOLD
    ws["E27"] = money(spec["_contract_amount"]); ws["E27"].font = BOLD

    # ---- the separate summary sheet: how we got to the contract amount -------
    s = wb.create_sheet("contract amount summary")
    s.column_dimensions["A"].width = 72
    s.column_dimensions["B"].width = 18
    s["A1"] = f"How we got to the contract amount — {pkg} — {spec['_subcontractor']}"
    s["A1"].font = Font(bold=True, size=12)
    s["A2"] = ("Reconciled to GMP R2 leveling sheet " + spec["_leveling_sheet"] +
               ". The leveling total is SUM of the priced column; values marked with "
               "an asterisk on that sheet are options and do not count toward it.")
    s["A2"].alignment = Alignment(wrap_text=True)
    s.row_dimensions[2].height = 30

    r = 4
    for label, amt in spec["_amount_buildup"]:
        s[f"A{r}"] = label
        s[f"B{r}"] = amt
        s[f"B{r}"].number_format = '#,##0.00;-#,##0.00'
        for c in (f"A{r}", f"B{r}"):
            s[c].border = BOX
        r += 1
    s[f"A{r}"] = "CONTRACT TOTAL"
    s[f"B{r}"] = spec["_contract_amount"]
    s[f"B{r}"].number_format = '#,##0.00'
    for c in (f"A{r}", f"B{r}"):
        s[c].font = BOLD
        s[c].border = BOX

    f = dest / f"{spec['file_stem']} - Att A Review Cover Sheet.xlsx"
    wb.save(f)
    total = sum(a for _, a in spec["_amount_buildup"])
    print(f"wrote {f.relative_to(ROOT)}")
    print(f"  buildup sums to {total:,.0f} — contract amount {spec['_contract_amount']:,.0f}"
          f"  {'MATCH' if round(total) == spec['_contract_amount'] else 'MISMATCH'}")
    return f


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else "RFP-008")
