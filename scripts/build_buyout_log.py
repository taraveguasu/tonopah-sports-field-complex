#!/usr/bin/env python3
"""
Build the buy-out log and the subcontract release priority list.

Two questions this answers:
  1. What is bought out, from whom, for how much, and where does each package
     stand -- the log.
  2. In what order do the subcontracts have to go out the door, and which one
     is already hurting us -- the priority list.

The ranking is NOT a scoring rubric. It is a backward pass through the
04.21.26 preliminary schedule:

    latest responsible release date (LRRD)
        = first field activity that needs this sub
        - procurement duration
        - submittal review duration
        - submittal preparation duration
        - execution allowance (bonds, insurance, signatures)

all in working days. A package whose LRRD is already behind us is late, and
how late is the rank. Cost, delegated design, coordination and risk break ties
and move a package up a tier, but they never invent urgency the schedule does
not support.

Durations come from the schedule where the schedule names the item. Where it
does not -- and it names nothing for turf, track, fencing, landscaping or any
of the sports-field products -- the duration is trade judgment and is labeled
"judgment" in the output so it can be argued with.

Data authority:
  - carried subcontractor + carried value: GMP R2 `BackSheet` (line 16-146),
    which is where the GMP's own money lands. NOT the bid tabulation, which is
    a transcription, and NOT awarded-sub-mapping.json, which predates the GMP.
  - schedule activities: 05.7 Preliminary Construction Schedule (04.21.26),
    read at activity level.
  - open risk items: 01-index/pm-open-items.json.

Usage:  python3 scripts/build_buyout_log.py
"""

import json
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
GMP = ROOT / "00-source-docs" / "GMP" / "NCSD THS Field - GMP R2 - 070126.xlsm"
OUT = ROOT / "04-output"
INDEX = ROOT / "01-index"

TODAY = date(2026, 8, 7)
# The schedule's own subcontract procurement window, 04.21.26 schedule ID 19.
PLANNED_BUYOUT = (date(2026, 7, 21), date(2026, 7, 27))
EXEC_DAYS = 15          # issue -> executed, incl. bonds and insurance certs
PM_REVIEW_DAYS = 10     # Attachment A draft -> PM review, revision, packet assembly

# Packages whose Attachment A has been drafted and sent to the PM for review.
# Add the package_id here when its exhibit lands in 02-drafts/.
ATT_A_DRAFTED = {"RFP-008", "ITB-072"}

# --------------------------------------------------------------------------
# Which GMP BackSheet lines roll into which bid package.
# The "(SEE LINE #xxx)" pointers in the BackSheet are the source for the
# rollups -- they are printed in the workbook, not inferred.
# --------------------------------------------------------------------------
ROLLUP = {
    "RFP-002": [2, 4],                    # 004 Building Wrecking -> 002
    "ITB-008": [7],
    "RFP-008": [8, 9, 14, 15],            # utilities, asphalt, site signage -> 008
    "RFP-016": [16],
    "ITB-018": [17],
    "ITB-019": [18, 19, 20],              # supply + track install + field install
    "RFP-021": [21],
    "RFP-022": [22],
    "RFP-023": [23],
    "RFP-030": [24, 30],                  # site concrete -> 030
    "RFP-031": [31],
    "RFP-033": [26, 33],                  # site metals -> 033
    "ITB-040": [40],
    "ITB-044": [44],                      # carried $0, "(SEE LINE #060)"
    "RFP-045": [45, 49],                  # metal panels -> 045
    "ITB-054": [54],
    "ITB-056": [56, 57],
    "RFP-060": [60],
    "ITB-062": [62],                      # carried $0, "(SEE LINE #060)"
    "ITB-066": [66],                      # removed by Owner VE
    "ITB-067": [67],
    "ITB-070": [70],
    "ITB-071": [71, 83],                  # manufactured specialties -> 071
    "ITB-072": [72],
    "ITB-074": [74, 75],                  # fire protection specialties -> 074
    "ITB-077": [77],
    "ITB-078": [78],
    "ITB-085": [85],
    "ITB-089": [89],
    "RFP-094": [94],
    "RFP-098": [98],
    "RFP-100": [100, 101, 102],           # controls + T&B -> 100
    "RFP-103": [103, 105, 106, 107],      # low voltage, security, field lighting -> 103
    "RFP-109": [109],
}

# --------------------------------------------------------------------------
# The backward pass inputs. One package can have several independent chains --
# a trade's rough-in material, its long-lead equipment, and its shop drawings
# often serve different field activities on different dates. The BINDING chain
# is whichever produces the earliest release date; the others are reported so
# the float behind them is visible.
#
# Each chain: (label, field activity, need date, submit, review, procure, basis)
#   "sched"    the 04.21.26 schedule names that activity and this is its duration
#   "judgment" the schedule is silent -- trade estimate, also flagged as a gap
#
# A chain with submit/review/procure all zero is a pure mobilization chain: the
# sub needs nothing but an executed agreement and notice.
# --------------------------------------------------------------------------
D = lambda y, m, d: date(y, m, d)

DRIVERS = {
    "ITB-008": dict(chains=[
        ("Survey control and staking", "Temp Fence & Mobilize", D(2026, 11, 24), 10, 5, 0, "judgment"),
    ], note=(
        "No submittal or procurement activity anywhere in the schedule. Control has "
        "to be set before the first field activity, so mobilization is the need date, "
        "not a later staking task.")),

    "RFP-002": dict(chains=[
        ("NESHAP/NDEP notification and abatement work plan", "Abatement ahead of Demo",
         D(2026, 11, 24), 15, 10, 0, "judgment"),
    ], note=(
        "A regulatory clock, not a material clock. NESHAP requires written notification "
        "before an abatement start, and NDEP/NV OSHA want the work plan and licenses "
        "first. The schedule carries no submittal line for abatement at all, and "
        "abatement has to finish before the 12/3/26 wrecking.")),

    "RFP-008": dict(chains=[
        ("Mobilization — first sub on the field", "Temp Fence & Mobilize",
         D(2026, 11, 24), 10, 0, 0, "judgment"),
        ("Underground utility materials", "Storm Drain", D(2027, 2, 8), 15, 15, 15, "sched"),
    ], note=(
        "Schedule IDs 34/54/71 give the utility chain 15/15/15 with procurement "
        "complete 9/29/26, which has float. Mobilization is what binds: this sub "
        "opens the site and every other site package sequences behind it.")),

    "RFP-030": dict(chains=[
        ("Footing rebar", "Retaining Walls / Build Building Pad", D(2027, 2, 8), 15, 15, 15, "sched"),
        ("Anchor bolt and embed coordination from RFP-033", "Foundations",
         D(2027, 2, 15), 15, 15, 15, "sched"),
    ], note=(
        "Schedule IDs 22/49/68 for rebar. The real constraint is not rebar, it is "
        "everything embedded in this sub's concrete: RFP-033 anchor bolts, ITB-019 "
        "equipment sleeves, RFP-094 bleacher anchors, ITB-078 flagpole foundations, "
        "and the trench drains ruled into RFP-030 on 07.31.26. Each of those suppliers "
        "has to be under contract early enough to hand over a layout.")),

    "RFP-103": dict(chains=[
        ("Electrical rough-in material", "UG Electrical & Musco Bases",
         D(2027, 2, 8), 15, 15, 15, "sched"),
        ("Musco light fixtures", "Install Musco Light Poles", D(2027, 3, 8), 10, 15, 50, "sched"),
        ("Electrical lighting package", "Set Electrical Equipment",
         D(2027, 7, 6), 40, 15, 140, "sched"),
    ], note=(
        "Three separate chains, IDs 29/52/70, 31/47/67 and 30/63/84, and the one that "
        "binds is the least obvious. Rough-in material for the 2/8/27 underground "
        "needs release by 11/13/26; the Musco poles by 11/2/26; but the electrical "
        "LIGHTING PACKAGE runs 40 d submit + 15 d review + 140 d procure, and 210 "
        "working days back from Set Electrical Equipment on 7/6/27 lands on 9/15/26. "
        "That 140-day procurement is the longest on the job and is what defines the "
        "5/3/27 Material Procurement Complete milestone -- the schedule's own last "
        "procurement date belongs to this package.")),

    "RFP-033": dict(chains=[
        ("Anchor bolts — embedded by RFP-030", "Foundations", D(2027, 2, 15), 15, 15, 15, "sched"),
        ("Structural steel", "Steel, Joists, & Deck Erection", D(2027, 5, 3), 40, 30, 80, "sched"),
        ("Joist & deck — deferred submittal", "Steel, Joists, & Deck Erection",
         D(2027, 5, 3), 40, 15, 100, "sched"),
    ], note=(
        "IDs 25/48/73 for anchor bolts, 24/60/87 for steel, 42/90/81 for joist and "
        "deck. JOIST & DECK BINDS: 40 d submit + 15 d review + 100 d procure = 170 "
        "working days back from erection on 5/3/27, the longest single chain on the "
        "job, and it is a deferred submittal so the review runs through the building "
        "official rather than stopping at the EOR. The steel itself is close behind "
        "at 165 days, carrying the longest submittal review on the job at 30 days "
        "(ID 60). And a third obligation lands far earlier than either: the ANCHOR "
        "BOLTS are needed 2/15/27 for foundations, two and a half months before "
        "erection, embedded in RFP-030's concrete.")),

    "RFP-094": dict(chains=[
        ("Anchor layout to RFP-030 — deferred structural submittal", "Bleacher Footings",
         D(2027, 3, 8), 40, 15, 0, "sched"),
        ("Bleacher and press box fabrication", "Bleacher Installation",
         D(2027, 4, 19), 40, 15, 40, "sched"),
    ], note=(
        "IDs 38/59/79. Two chains that look similar and are not: the product ships in "
        "time for a 4/19/27 install, but the ANCHOR LAYOUT has to reach RFP-030 by "
        "3/8/27 for the footings, and it comes out of a deferred submittal that goes "
        "to the AHJ, not just the EOR.")),

    "RFP-031": dict(chains=[
        ("CMU", "Erect CMU", D(2027, 4, 5), 15, 15, 40, "sched"),
        ("Masonry rebar", "Erect CMU", D(2027, 4, 5), 15, 15, 15, "sched"),
    ], note=(
        "IDs 36/55/72 and 23/50/69. Hollow metal frames from ITB-056 are set in this "
        "masonry, so ITB-056's frame chain has to beat the same 4/5/27 date.")),

    "ITB-056": dict(chains=[
        ("Hollow metal frames — set in CMU", "Erect CMU", D(2027, 4, 5), 20, 15, 50, "sched"),
        ("Doors and hardware", "Doors & Hardware", D(2027, 8, 17), 40, 20, 100, "sched"),
    ], note=(
        "IDs 21/56/77 for frames, 26/61/86 for doors. Buying this package to the "
        "8/17/27 door install misses by five months -- the FRAMES are what bind, and "
        "they have to be on site for masonry.")),

    "RFP-022": dict(chains=[
        ("Made-to-order turf system", "Sports Field Base", D(2027, 5, 3), 20, 15, 70, "judgment"),
    ], note=(
        "SCHEDULE GAP. ID 39 submits Sports Field Product Data (20 d) and then nothing "
        "-- there is no procurement activity for $1,139,000 of turf. Matrix Helix 46 oz "
        "with Elia infill over a SOTERIA 20mm pad is manufactured to order; 70 working "
        "days after approval is the estimate used here and it should be confirmed with "
        "the manufacturer before this ranking is relied on.")),

    "RFP-098": dict(chains=[
        ("Plumbing rough-in material", "Rough-in Underground Plumbing",
         D(2027, 3, 1), 15, 15, 15, "sched"),
        ("Plumbing equipment and fixtures", "MEP Trim", D(2027, 7, 20), 40, 15, 60, "sched"),
    ], note="IDs 32/53/75 and 33/64/85. The underground rough-in is what binds."),

    "RFP-045": dict(chains=[
        ("Roof materials", "Roof, Fascia, Soffit Panels", D(2027, 6, 8), 20, 15, 100, "sched"),
        ("Metal panels", "Roof, Fascia, Soffit Panels", D(2027, 6, 8), 30, 15, 100, "sched"),
    ], note=(
        "IDs 37/57/76 and 35/58/78. Two 100-day procurements against the same field "
        "date. Berridge C-Lock per the 08.06.26 B1 ruling -- that ruling reverses a "
        "written architect's RFI answer and has to be settled before release or the "
        "sub orders the wrong panel and the 100 days start over.")),

    "ITB-019": dict(chains=[
        ("Equipment sleeves and anchors — embedded by RFP-030", "Sports Equipment Sleeves",
         D(2027, 4, 26), 20, 15, 40, "judgment"),
        ("Athletic equipment", "Sports Field Equipment Install", D(2027, 7, 13), 20, 15, 40, "judgment"),
    ], note=(
        "SCHEDULE GAP -- no procurement activity for athletic equipment. The equipment "
        "installs 7/13/27, but the SLEEVES are embedded in RFP-030's concrete on "
        "4/26/27. Buying this package to the install date strands the concrete sub.")),

    "RFP-100": dict(chains=[
        ("Mechanical rough-in material", "MEP Overhead Rough", D(2027, 5, 24), 15, 15, 15, "sched"),
        ("Mechanical equipment", "Set Mechanical Equipment", D(2027, 7, 13), 40, 15, 80, "sched"),
    ], note=(
        "IDs 27/51/74 and 28/62/83. Test & Balance sits inside this package per the "
        "08.06.26 A7 ruling and runs 8/24-9/7/27, after the field is complete.")),

    "RFP-023": dict(chains=[
        ("Fence and gate fabrication", "Fence Posts", D(2027, 3, 8), 15, 15, 30, "judgment"),
    ], note=(
        "SCHEDULE GAP -- no submittal or procurement activity for fencing. This is "
        "also the only package whose scope-review record on file is a blank template, "
        "so there is no evidence the scope review was ever held.")),

    "RFP-016": dict(chains=[
        ("Irrigation mainline sleeving ahead of curb and paving", "Concrete Curbing",
         D(2027, 3, 8), 15, 15, 20, "judgment"),
        ("Planting and turf-area finish materials", "Sports Field Base",
         D(2027, 5, 3), 15, 15, 20, "judgment"),
    ], note=(
        "SCHEDULE GAP -- landscaping and irrigation appear in no procurement and no "
        "field activity. Sleeving is the early obligation; planting is late. This "
        "package also carries the shot put and discus mix ruled to RFP-016 on "
        "08.06.26 (A5), which RFP-030 and ITB-019 have to coordinate around.")),

    "ITB-089": dict(chains=[
        ("Deferred structural submittal and fabrication", "Install Musco Light Poles",
         D(2027, 3, 8), 30, 15, 40, "judgment"),
    ], note=(
        "SCHEDULE GAP -- the scoreboard has no procurement activity. Sheet G0-00 lists "
        "it as a deferred submittal, so it carries delegated design through the AHJ. "
        "Its foundation is RFP-030's pour and its feed is RFP-103's conduit, so both "
        "of those need its layout.")),

    "RFP-021": dict(chains=[
        ("Track surfacing materials", "Running Track Base", D(2027, 6, 1), 20, 15, 40, "judgment"),
    ], note=(
        "SCHEDULE GAP -- no procurement line for track surfacing. Surfacing runs "
        "7/13-7/26/27 and Sports Field Substantial Completion is 7/26/27: zero float "
        "to a contractual milestone, on a temperature-dependent application in a "
        "high-desert July, with only two real bidders behind it.")),

    "RFP-060": dict(chains=[
        ("Framing and sheathing materials", "Exterior Framing & Sheathing",
         D(2027, 5, 24), 15, 15, 20, "judgment"),
    ], note=(
        "One agreement carrying four exhibits per the 08.06.26 D ruling -- RFP-060 plus "
        "ITB-044, ITB-062 and ITB-077. None of the four can issue until all four are "
        "drafted, so this release date is really the earliest of that whole group.")),

    "RFP-109": dict(chains=[
        ("Prefabricated building, factory built to order", "Press Box Installation window",
         D(2027, 6, 1), 20, 15, 50, "judgment"),
    ], note=(
        "SCHEDULE GAP -- the ticket booth appears in no procurement and no field "
        "activity. It is a manufactured building (Porta-King DURASTEEL PC64 per the "
        "08.04.26 ruling), not a delivered product, and its pad is RFP-030's.")),

    "ITB-085": dict(chains=[
        ("Approved equipment cut sheets set the MEP rough-in", "MEP Rough In Walls",
         D(2027, 6, 8), 20, 15, 0, "judgment"),
        ("Equipment delivery", "Casework", D(2027, 7, 13), 20, 15, 40, "judgment"),
    ], note=(
        "SCHEDULE GAP. The equipment lands late, but RFP-098/100/103 cannot rough in "
        "walls without approved cut sheets -- the submittal is the early obligation, "
        "not the delivery.")),

    "ITB-054": dict(chains=[
        ("Coiling door fabrication", "Windows & Coiling Doors", D(2027, 6, 15), 20, 15, 45, "judgment"),
    ], note=(
        "SCHEDULE GAP -- special doors have no procurement activity. The Door 106 "
        "panic hardware ambiguity is shared with ITB-056 and is a life-safety item.")),

    "ITB-018": dict(chains=[
        ("Site furnishings supply for ITB-019 to install", "Sports Field Equipment Install",
         D(2027, 7, 13), 15, 15, 30, "judgment"),
    ], note=(
        "Supply only per the 08.06.26 A2 ruling -- ITB-019 installs. Same vendor as "
        "ITB-019, so both exhibits should travel together.")),

    "ITB-074": dict(chains=[
        ("In-wall backing coordination with RFP-060", "Framing", D(2027, 6, 1), 15, 15, 0, "judgment"),
        ("Specialties and fire protection cabinets", "MEP Trim", D(2027, 7, 20), 15, 15, 30, "judgment"),
    ], note=(
        "The accessories install late but their backing goes in during framing, so the "
        "layout is the early obligation. Carries the fire protection cabinets and the "
        "wall and door protection per the 08.06.26 D and A3 rulings.")),

    "ITB-040": dict(chains=[
        ("Sealant and weather barrier materials", "Roof, Fascia, Soffit Panels",
         D(2027, 6, 8), 15, 10, 15, "judgment"),
    ], note=(
        "Below-grade waterproofing only per the 07.31.26 ruling. 07 25 00 Weather "
        "Barriers is shared with RFP-045 along that same boundary.")),

    "ITB-078": dict(chains=[
        ("Foundation and sleeve detail to RFP-030", "Concrete Curbing",
         D(2027, 3, 8), 15, 15, 0, "judgment"),
        ("Flagpole delivery", "Exterior Finishes Complete", D(2027, 7, 19), 15, 15, 30, "judgment"),
    ], note=(
        "Small money, early dependency: the foundation is RFP-030's pour and needs "
        "this supplier's base detail long before the pole shows up.")),

    "ITB-077": dict(chains=[
        ("Locker fabrication — ASI Pro Collection", "Doors & Hardware",
         D(2027, 8, 17), 20, 15, 40, "judgment"),
    ], note=(
        "Inside the Jetstream agreement with RFP-060. No spec section was ever issued; "
        "the basis of design is Clarification No. 1 RFI #3.")),

    "ITB-044": dict(chains=[
        ("Insulation materials", "Insulation", D(2027, 6, 22), 10, 10, 15, "judgment"),
    ], note=(
        "Carried at $0 in the GMP -- '(SEE LINE #060)'. No separate money; releases "
        "inside the Jetstream agreement.")),

    "ITB-062": dict(chains=[
        ("Ceiling grid and tile", "MEP Trim", D(2027, 7, 20), 15, 15, 30, "judgment"),
    ], note=(
        "Carried at $0 in the GMP -- '(SEE LINE #060)'. No separate money; releases "
        "inside the Jetstream agreement.")),

    "ITB-067": dict(chains=[
        ("Sealed concrete materials", "Flooring", D(2027, 8, 3), 10, 10, 15, "judgment"),
    ], note=(
        "Scope grew when ITB-066 was VE'd out -- those locations became sealed concrete "
        "under this package per the 08.04.26 ruling.")),

    "ITB-071": dict(chains=[
        ("Boards and display case fabrication", "Casework", D(2027, 7, 13), 15, 10, 25, "judgment"),
    ], note=(
        "The 'ADD Menu Display Case' alternate is worded identically in ITB-072 and "
        "ITB-085, and this package's own title names it -- three packages may be "
        "touching one item.")),

    "ITB-072": dict(chains=[
        ("Sign fabrication", "Exterior Finishes Complete", D(2027, 7, 19), 15, 10, 30, "judgment"),
    ], note="Same vendor as ITB-078 and ITB-089 (YESCO) -- three exhibits, one contract negotiation."),

    "ITB-070": dict(chains=[
        ("Labor only", "Construction Clean", D(2027, 8, 24), 5, 5, 0, "judgment"),
    ], note=(
        "No scope narrative was ever issued (08.06.26 E1 ruling) -- the exhibit has to "
        "be drafted from the four proposals on hand.")),

    "ITB-066": dict(chains=[], note=(
        "Scope removed by the Owner in a value engineering exercise, 08.04.26. No "
        "subcontract will be issued. Those locations became sealed concrete under "
        "ITB-067.")),
}


# Non-schedule factors. These do not create urgency; they raise a tier and they
# are what the PM actually has to resolve before the exhibit can issue.
FACTORS = {
    "RFP-103": ["Largest subcontract on the job at $1,074,541",
                "Longest procurement on the job (140 d) and it defines Material Procurement Complete",
                "Delegated design: Musco pole foundations and sports lighting photometrics",
                "Single-source manufacturer (Musco) — no recovery by rebidding",
                "OPEN: the ASI formalizing the no-fire-alarm direction has not issued; "
                "the GMP Exclusion covers the money but not the drawings"],
    "RFP-030": ["Second-largest subcontract at $1,278,460",
                "Coordination hub — five other packages embed work in this concrete",
                "Carries the 07.31.26 rulings on trench drains and athletic equipment footings"],
    "RFP-008": ["First sub on the field; every site package sequences behind it",
                "OPEN: package title carries site signage and striping the narrative never "
                "scopes and the sub excluded twice",
                "OPEN: accepted VE #10 at −$16,750 has no matching line in the sub's own homework",
                "OPEN: stockpiling — narrative prohibits, GMP assumes, VE #10 relies on it"],
    "RFP-002": ["Regulatory lead, not material lead — notification periods cannot be compressed",
                "Asbestos: an abatement failure stops the demolition that stops everything else",
                "OPEN: the ES campus demolition alternate is worded identically in RFP-002 "
                "and RFP-008 and it is still unresolved whether it splits by trade"],
    "RFP-033": ["Longest submittal review on the job (30 d) with no slack behind it",
                "Delegated design: joist & deck is a deferred submittal",
                "OPEN: goal post structural design is listed as a deferred submittal on G0-00 "
                "and is assigned to no package"],
    "RFP-094": ["Deferred submittal — delegated design through the AHJ",
                "OPEN: the only specifications that exist are the written blocks on sheet A1-40",
                "Anchor layout feeds RFP-030's footings a month before fabrication would suggest"],
    "RFP-022": ["$1,139,000 with no procurement activity anywhere in the schedule",
                "Made-to-order product; a late release cannot be recovered by adding crews",
                "OPEN: Sprinturf's bid never split between RFP-016 and RFP-022"],
    "RFP-021": ["ZERO float — surfacing finishes the same day the field is contractually complete",
                "Weather-dependent application in July at 6,000 ft",
                "Only two real bidders; no recovery by rebidding"],
    "RFP-016": ["OPEN AND BLOCKING: the GMP carries Black Canyon at $350,200 but their "
                "proposal on file states $111,275 — a 3.1x delta that must be reconciled "
                "before this issues"],
    "ITB-056": ["Frames, not doors, are the driver — the buy-out date is CMU, not door install",
                "OPEN: Door 106 panic hardware ambiguity, shared with ITB-054, life-safety"],
    "ITB-019": ["Sleeves are embedded by another sub four months before this sub installs",
                "OPEN: 12 93 00 is claimed by both ITB-018 and ITB-019",
                "Exerplay carries four GMP lines across two packages — they must split cleanly"],
    "RFP-060": ["One agreement, four exhibits (RFP-060, ITB-044, ITB-062, ITB-077) — "
                "none can issue until all four are drafted",
                "ITB-044 and ITB-062 are carried at $0; their money is inside RFP-060's $266,211"],
    "RFP-023": ["OPEN: the scope review agenda on file is a blank template — no evidence "
                "the scope review was ever held"],
    "RFP-045": ["OPEN: the 08.06.26 B1 ruling puts the GMP over RFI #23, reversing a written "
                "architect's answer — settle before release or the sub buys the wrong panel"],
    "RFP-109": ["Manufactured building with no schedule activity of any kind"],
    "ITB-089": ["Deferred submittal with delegated design and no procurement activity"],
    "ITB-070": ["No scope narrative was ever issued — the exhibit is drafted from proposals"],
    "ITB-066": ["Scope removed by the Owner. No subcontract."],
}

TITLES = {
    "RFP-002": "Abatement & Building Wrecking",
    "RFP-008": "Site Demolition, Salvage, Earthwork, Asphalt Paving, Wet Utilities, & Site Signage & Striping",
    "RFP-016": "Landscaping & Irrigation",
    "RFP-021": "Running Track Surfacing",
    "RFP-022": "Synthetic Turf Sports Field",
    "RFP-023": "Fencing & Gates",
    "RFP-030": "Concrete",
    "RFP-031": "Masonry",
    "RFP-033": "Structural Steel & Ornamental Metals",
    "RFP-045": "Metal Roofing, Fascia & Soffit Panels",
    "RFP-060": "Framing, Drywall & Painting (incl. FRP)",
    "RFP-094": "Bleachers & Press Box",
    "RFP-098": "Plumbing Systems",
    "RFP-100": "HVAC & Building Control Systems (incl. Test & Balance)",
    "RFP-103": "Electrical & Low Voltage Systems (incl. Sports Field Lighting)",
    "RFP-109": "Prefabricated Ticket Booth",
    "ITB-008": "Surveying, Layout & Staking",
    "ITB-018": "Site Furnishings",
    "ITB-019": "Track & Field Athletic Equipment",
    "ITB-040": "Moisture Protection & Sealants (incl. Acoustical Caulking)",
    "ITB-044": "Insulation",
    "ITB-054": "Special Doors",
    "ITB-056": "Doors, Frames & Hardware",
    "ITB-062": "Acoustical Ceiling Treatments",
    "ITB-066": "Fluid-Applied Flooring",
    "ITB-067": "Concrete Finishing",
    "ITB-070": "Final Cleaning",
    "ITB-071": "Visual Display Boards & Menu Display Case",
    "ITB-072": "Building Signage",
    "ITB-074": "Building & Fire Protection Specialties",
    "ITB-077": "Lockers",
    "ITB-078": "Flagpoles",
    "ITB-085": "Warming Kitchen Food Service Equipment",
    "ITB-089": "Scoreboards",
}

# Packages carried at $0 in the GMP because their money sits in another line.
# The BackSheet names no firm on those rows, but the subcontractor is known --
# it is whoever holds the line the money was rolled into.
CARRIED_INSIDE = {
    "ITB-044": "Jetstream Construction",     # money inside #060
    "ITB-062": "Jetstream Construction",     # money inside #060
}

# Groups. "agreement" is a PM ruling and IS applied to the dates -- every member
# of a single agreement releases on the earliest member's date, because one
# agreement cannot issue in pieces. "vendor" is the same firm on several
# packages and is only REPORTED: whether to write one agreement or several is
# the PM's call, so it is not baked into the ranking.
GROUPS = [
    ("agreement", "Jetstream — one agreement, four exhibits (PM ruling 08.06.26 D)",
     ["RFP-060", "ITB-044", "ITB-062", "ITB-077"]),
    ("vendor", "Exerplay — supplies ITB-018, supplies and installs ITB-019",
     ["ITB-018", "ITB-019"]),
    ("vendor", "YESCO — signage, flagpoles and scoreboard",
     ["ITB-072", "ITB-078", "ITB-089"]),
    ("vendor", "US Mechanical — plumbing and HVAC",
     ["RFP-098", "RFP-100"]),
]

# Packages sharing one subcontract agreement or one vendor.
COMBINED = {
    "RFP-060": "One agreement with ITB-044, ITB-062, ITB-077 (Jetstream)",
    "ITB-044": "Inside the RFP-060 agreement (Jetstream)",
    "ITB-062": "Inside the RFP-060 agreement (Jetstream)",
    "ITB-077": "Inside the RFP-060 agreement (Jetstream)",
    "ITB-018": "Same vendor as ITB-019 (Exerplay) — release together",
    "ITB-019": "Same vendor as ITB-018 (Exerplay) — release together",
    "ITB-072": "Same vendor as ITB-078, ITB-089 (YESCO)",
    "ITB-078": "Same vendor as ITB-072, ITB-089 (YESCO)",
    "ITB-089": "Same vendor as ITB-072, ITB-072 (YESCO)",
    "RFP-098": "Same vendor as RFP-100 (US Mechanical)",
    "RFP-100": "Same vendor as RFP-098 (US Mechanical)",
}


def busoff(d, days):
    """Move `days` working days from date d. Negative goes backward.

    Mon-Fri only. Holidays are not modeled -- the schedule this is derived from
    does not publish its calendar, so adding a holiday list would be inventing
    precision. Every date here is therefore slightly optimistic.
    """
    step = 1 if days >= 0 else -1
    n = abs(days)
    while d.weekday() >= 5:
        d += timedelta(days=step)
    while n:
        d += timedelta(days=step)
        while d.weekday() >= 5:
            d += timedelta(days=step)
        n -= 1
    return d


def busdiff(a, b):
    """Working days from a to b (positive if b is later)."""
    lo, hi, sign = (a, b, 1) if a <= b else (b, a, -1)
    n, d = 0, lo
    while d < hi:
        d += timedelta(days=1)
        if d.weekday() < 5:
            n += 1
    return n * sign


def read_backsheet():
    """Carried subcontractor and carried value, per GMP scope line."""
    wb = openpyxl.load_workbook(GMP, data_only=True)
    ws = wb["BackSheet"]
    lines = {}
    for r in range(16, 150):
        num, desc, firm, amt = (ws.cell(r, c).value for c in (4, 5, 6, 7))
        # BackSheet columns shift by one relative to the raw dump; find them.
        num, desc, firm, amt = (ws.cell(r, 4).value, ws.cell(r, 5).value,
                                ws.cell(r, 6).value, ws.cell(r, 7).value)
        if isinstance(num, (int, float)) and isinstance(desc, str):
            lines[int(num)] = dict(scope=desc.strip(), firm=firm,
                                   amount=amt if isinstance(amt, (int, float)) else 0)
    return lines


def open_items_by_package():
    data = json.loads((INDEX / "pm-open-items.json").read_text())
    out = {}
    for it in data["items"]:
        for pkg in it.get("packages", []) or []:
            out.setdefault(pkg, []).append(
                f"[{it.get('severity', '?').upper()}] {it.get('id', '')} {it.get('title', '')}")
    return out


def build():
    lines = read_backsheet()
    opens = open_items_by_package()
    rows = []

    for pkg, drv in DRIVERS.items():
        nums = ROLLUP[pkg]
        # The carried firm sits on whichever rolled line actually holds money.
        firm, amount, carried_on = None, 0.0, []
        for n in nums:
            ln = lines.get(n, {})
            amt = ln.get("amount") or 0
            carried_on.append(f"#{n:03d} {ln.get('scope', '?')}"
                              + (f" — {ln['firm']}" if ln.get("firm") else "")
                              + (f" ${amt:,.0f}" if amt else " $0"))
            if amt:
                amount += amt
                if firm is None and isinstance(ln.get("firm"), str) \
                        and not ln["firm"].startswith("("):
                    firm = ln["firm"]
        if firm is None and pkg in CARRIED_INSIDE:
            firm = CARRIED_INSIDE[pkg]
        if firm is None:
            for n in nums:
                f = lines.get(n, {}).get("firm")
                if isinstance(f, str) and not f.startswith("("):
                    firm = f
                    break

        # Every chain gets its own release date; the earliest one binds and the
        # rest are reported so the float behind them is visible rather than
        # assumed.
        evaluated = []
        for label, act, need, sub, rev, pro, basis in drv["chains"]:
            total = sub + rev + pro + EXEC_DAYS
            rel = busoff(need, -total)
            evaluated.append(dict(chain=label, field_activity=act,
                                  need_date=need.isoformat(),
                                  submit_days=sub, review_days=rev,
                                  procure_days=pro, exec_days=EXEC_DAYS,
                                  chain_days=total, basis=basis,
                                  release_by=rel.isoformat(),
                                  working_days_late=busdiff(rel, TODAY)))
        evaluated.sort(key=lambda c: c["release_by"])
        binding = evaluated[0] if evaluated else None

        rows.append(dict(
            package_id=pkg,
            source="1% (NRS 338.16995)" if pkg.startswith("RFP") else "Non-1% (ITB)",
            title=TITLES[pkg],
            carried_sub=firm or "—",
            carried_value=round(amount, 2),
            gmp_lines="; ".join(carried_on),
            chains=evaluated,
            binding_chain=binding["chain"] if binding else "",
            first_field_activity=binding["field_activity"] if binding else "— not released",
            field_need_date=binding["need_date"] if binding else "",
            submit_days=binding["submit_days"] if binding else 0,
            review_days=binding["review_days"] if binding else 0,
            procure_days=binding["procure_days"] if binding else 0,
            exec_days=EXEC_DAYS,
            duration_basis=binding["basis"] if binding else "n/a",
            chain_days=binding["chain_days"] if binding else 0,
            latest_release_date=binding["release_by"] if binding else "",
            working_days_late=binding["working_days_late"] if binding else None,
            float_vs_planned_buyout=(
                busdiff(PLANNED_BUYOUT[1], date.fromisoformat(binding["release_by"]))
                if binding else None),
            draft_by=(busoff(date.fromisoformat(binding["release_by"]), -PM_REVIEW_DAYS)
                      .isoformat() if binding else ""),
            schedule_note=drv["note"],
            other_factors=FACTORS.get(pkg, []),
            open_items=opens.get(pkg, []),
            combined=COMBINED.get(pkg, ""),
            att_a_status=("DRAFTED — with PM" if pkg in ATT_A_DRAFTED
                          else "not started"),
            subcontract_status="not issued",
        ))

    # Apply agreement groups BEFORE ranking: an agreement cannot issue in pieces,
    # so every member inherits the earliest member's release date. Vendor groups
    # are recorded but not applied -- that is a PM decision, not an arithmetic one.
    by_id = {r["package_id"]: r for r in rows}
    for kind, label, members in GROUPS:
        present = [by_id[m] for m in members if m in by_id and by_id[m]["latest_release_date"]]
        if not present:
            continue
        earliest = min(p["latest_release_date"] for p in present)
        driver = next(p["package_id"] for p in present if p["latest_release_date"] == earliest)
        for p in present:
            p["group_kind"] = kind
            p["group"] = label
            p["group_release_by"] = earliest
            p["group_driver"] = driver
            if kind == "agreement" and p["latest_release_date"] != earliest:
                p["own_release_by"] = p["latest_release_date"]
                p["latest_release_date"] = earliest
                p["draft_by"] = busoff(date.fromisoformat(earliest), -PM_REVIEW_DAYS).isoformat()
                p["working_days_late"] = busdiff(date.fromisoformat(earliest), TODAY)
                p["schedule_note"] += (
                    f"  PULLED FORWARD from {p['own_release_by']} to {earliest}: this "
                    f"exhibit is part of {label}, and the agreement cannot issue until "
                    f"every exhibit attached to it is finished.")

    # Rank: late first (most late first), then by earliest latest-release-date.
    active = [r for r in rows if r["latest_release_date"]]
    removed = [r for r in rows if not r["latest_release_date"]]
    active.sort(key=lambda r: (r["latest_release_date"], -r["carried_value"]))
    for i, r in enumerate(active, 1):
        r["priority_rank"] = i
        # Slack = working days remaining before the latest responsible release
        # date. Negative means that date has already passed.
        slack = -r["working_days_late"]
        r["working_days_slack"] = slack
        r["draft_slack"] = busdiff(TODAY, date.fromisoformat(r["draft_by"]))
        r["tier"] = ("1 — PAST DUE" if slack < 0 else
                     "1 — RELEASE NOW" if slack <= 40 else
                     "2 — next 60 work days" if slack <= 60 else
                     "3 — next 90 work days" if slack <= 90 else
                     "4 — scheduled")
    for r in removed:
        r["priority_rank"] = None
        r["working_days_slack"] = None
        r["draft_slack"] = None
        r["draft_by"] = ""
        r["tier"] = "— not released (scope removed)"

    out = active + removed
    (INDEX / "buyout-log.json").write_text(json.dumps(
        dict(_generated=TODAY.isoformat(),
             _authority="GMP R2 BackSheet for sub + value; 04.21.26 preliminary "
                        "schedule for dates; pm-open-items.json for risk",
             _planned_buyout_window=[d.isoformat() for d in PLANNED_BUYOUT],
             _total_carried=round(sum(r["carried_value"] for r in out), 2),
             packages=out), indent=1))
    return out


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------
BOLD = Font(bold=True)
WHITE = Font(bold=True, color="FFFFFF")
HDR = PatternFill("solid", fgColor="1F3864")
TIER = {"1": PatternFill("solid", fgColor="FFC7CE"),
        "2": PatternFill("solid", fgColor="FFEB9C"),
        "3": PatternFill("solid", fgColor="DDEBF7"),
        "4": PatternFill("solid", fgColor="E2EFDA")}
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def sheet(wb, name, headers, widths, rows, freeze="A2"):
    ws = wb.create_sheet(name)
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = WHITE
        c.fill = HDR
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    for r, row in enumerate(rows, 2):
        for i, v in enumerate(row, 1):
            c = ws.cell(r, i, v)
            c.alignment = WRAP
            c.border = BOX
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


def money_col(ws, col, nrows):
    for r in range(2, nrows + 2):
        ws.cell(r, col).number_format = '"$"#,##0'


def write_workbook(rows):
    OUT.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- 1. Release priority -------------------------------------------
    hdr = ["Rank", "Tier", "Package", "Title", "Carried Subcontractor",
           "Carried Value", "Binding Long-Lead Chain", "First Field Activity Needing This Sub",
           "Field Need Date", "Sub\nmit", "Rev\niew", "Proc\nure", "Exec", "Total",
           "Basis", "Attachment A\nDraft By", "Latest\nRelease Date", "Work Days\nof Slack", "Why This Rank"]
    w = [6, 22, 10, 30, 24, 14, 34, 26, 12, 6, 6, 6, 6, 7, 10, 12, 13, 9, 62]
    data = []
    for r in rows:
        why = r["schedule_note"]
        if r["other_factors"]:
            why += "\n\n" + "\n".join("• " + f for f in r["other_factors"])
        data.append([r["priority_rank"], r["tier"], r["package_id"], r["title"],
                     r["carried_sub"], r["carried_value"], r["binding_chain"],
                     r["first_field_activity"], r["field_need_date"],
                     r["submit_days"], r["review_days"], r["procure_days"],
                     r["exec_days"], r["chain_days"], r["duration_basis"], r["draft_by"],
                     r["latest_release_date"], r["working_days_slack"], why])
    ws = sheet(wb, "release priority", hdr, w, data)
    money_col(ws, 6, len(data))
    for i, r in enumerate(rows, 2):
        t = str(r["tier"])[0]
        if t in TIER:
            for c in range(1, 3):
                ws.cell(i, c).fill = TIER[t]
        ws.row_dimensions[i].height = 108

    # ---- 2. Buy-out log -------------------------------------------------
    hdr = ["Package", "List", "Title", "Carried Subcontractor", "Carried Value",
           "GMP Lines Carried (BackSheet)", "Group (agreement or shared vendor)", "Group Releases By",
           "Attachment A", "Cover Sheet", "Subcontract", "Executed", "NTP",
           "Attachment A Draft By", "Latest Release Date", "Field Need Date", "Rank"]
    w = [10, 20, 30, 24, 14, 46, 34, 14, 20, 14, 14, 12, 12, 14, 13, 12, 6]
    data = []
    for r in rows:
        data.append([r["package_id"], r["source"], r["title"], r["carried_sub"],
                     r["carried_value"], r["gmp_lines"], r.get("group", r["combined"]), r.get("group_release_by", ""),
                     r["att_a_status"],
                     "BUILT" if r["package_id"] == "RFP-008" else "",
                     r["subcontract_status"], "", "",
                     r["draft_by"], r["latest_release_date"], r["field_need_date"],
                     r["priority_rank"]])
    ws = sheet(wb, "buyout log", hdr, w, data)
    money_col(ws, 5, len(data))
    tot = len(data) + 2
    ws.cell(tot, 4, "TOTAL CARRIED").font = BOLD
    ws.cell(tot, 5, sum(r["carried_value"] for r in rows)).font = BOLD
    ws.cell(tot, 5).number_format = '"$"#,##0'
    for i in range(2, len(data) + 2):
        ws.row_dimensions[i].height = 58

    # ---- 3. Every long-lead chain, not just the binding one --------------
    hdr = ["Rank", "Package", "Binding?", "Long-Lead Chain", "Field Activity It Serves",
           "Field Need Date", "Submit", "Review", "Procure", "Exec", "Total",
           "Duration Basis", "Release By", "Work Days of Slack"]
    w = [6, 10, 9, 40, 30, 12, 8, 8, 8, 7, 7, 12, 12, 9]
    data = []
    for r in rows:
        for i, c in enumerate(r["chains"]):
            data.append([r["priority_rank"], r["package_id"],
                         "BINDING" if i == 0 else "float",
                         c["chain"], c["field_activity"], c["need_date"],
                         c["submit_days"], c["review_days"], c["procure_days"],
                         c["exec_days"], c["chain_days"], c["basis"],
                         c["release_by"], -c["working_days_late"]])
    ws = sheet(wb, "long-lead chains", hdr, w, data)
    for i in range(2, len(data) + 2):
        ws.row_dimensions[i].height = 30

    # ---- 4. Open items blocking release ---------------------------------
    hdr = ["Rank", "Package", "Carried Subcontractor", "Open Item / Risk", "Source"]
    w = [6, 10, 24, 88, 26]
    data = []
    for r in rows:
        for f in r["other_factors"]:
            data.append([r["priority_rank"], r["package_id"], r["carried_sub"], f,
                         "buy-out analysis"])
        for o in r["open_items"]:
            data.append([r["priority_rank"], r["package_id"], r["carried_sub"], o,
                         "pm-open-items.json"])
    ws = sheet(wb, "blocking items", hdr, w, data)
    for i in range(2, len(data) + 2):
        ws.row_dimensions[i].height = 30

    # ---- 5. Basis ---------------------------------------------------------
    ws = wb.create_sheet("basis")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 118
    basis = [
        ("Generated", TODAY.isoformat()),
        ("Sub + value authority", "GMP R2 BackSheet (NCSD THS Field - GMP R2 - 070126.xlsm), "
                                  "rows 16-146. This is where the GMP's money actually lands. NOT the "
                                  "05.12.26 bid tabulation, which is a transcription and is known to "
                                  "reverse RFP-008's two low rows."),
        ("Schedule authority", "05.7 Preliminary Construction Schedule (04.21.26), read at activity "
                               "level, cross-checked against the 05.29.26 GMP schedule milestones."),
        ("Planned buy-out window", "Schedule ID 19 'Procure Subcontractor Contracts', 5 d, "
                                   "Tue 7/21/26 – Mon 7/27/26. That window closed 11 days before this "
                                   "log was generated, with nothing issued. Every package therefore "
                                   "starts from a recovery position."),
        ("Ranking method", "Backward pass, working days: latest release date = first field activity "
                           "needing the sub − procurement − submittal review − submittal preparation "
                           "− 15 d execution allowance (bonds, insurance, signatures). Rank is by "
                           "that date, earliest first. Cost, delegated design, coordination and risk "
                           "move a package up a tier and are recorded in 'Why This Rank' — they never "
                           "create urgency the schedule does not support."),
        ("Duration basis 'sched'", "The 04.21.26 schedule names that submittal, review or procurement "
                                   "activity and this is its own duration."),
        ("Duration basis 'judgment'", "The schedule is silent for that package. The duration is a trade "
                                      "estimate and every one is also listed below as a schedule gap."),
        ("Execution allowance", f"{EXEC_DAYS} working days from issue-for-signature to executed, "
                                "covering the subcontractor's signature, payment and performance "
                                "bonds, and insurance certificates."),
        ("", ""),
        ("SCHEDULE GAPS FOUND", "The 04.21.26 schedule has no submittal or procurement activity for "
                                "ANY of the sports-field products, which is $3.1M of carried value: "
                                "RFP-022 synthetic turf ($1,139,000), RFP-021 track surfacing "
                                "($450,741), RFP-023 fencing ($240,789), RFP-016 landscape & "
                                "irrigation ($350,200), ITB-019 track & field equipment ($307,090), "
                                "ITB-089 scoreboard ($99,483), RFP-109 ticket booth ($27,019). Their "
                                "release dates below are computed from trade judgment."),
        ("STALE SCHEDULE LINES", "The schedule still procures fire alarm (80 d) and fire sprinkler "
                                 "(20 d) material and submits drawings for both. Fire alarm is an "
                                 "express GMP Exclusion (08.04.26 ruling) and fire suppression is "
                                 "carried N/A at $0. Neither is a package and neither should stay on "
                                 "a procurement critical path."),
        ("ROLLED-UP LINES", "Eight GMP lines are carried at $0 with a '(SEE LINE #xxx)' pointer and "
                            "are bought inside another package: #004 building wrecking → 002; "
                            "#009 utilities and #015 site signage → 008; #024 site concrete → 030; "
                            "#026 site metals → 033; #049 metal panels → 045; #044 insulation and "
                            "#062 acoustical ceilings → 060; #075 fire protection specialties → 074; "
                            "#102 test & balance → 100; #107 sports field lighting → 103. They are "
                            "shown on the log so none of them gets bought twice."),
        ("NOT IN THIS LOG", "CORE self-perform and general conditions (GEN1 temporary construction "
                            "requirements $57,630, GEN2 equipment $58,000, GEN3 waste management & "
                            "cleaning $83,545, #104 temporary power & lighting $2,550), contingency, "
                            "allowances, insurance and bonds. Per direction, this log covers the "
                            "bid packages only."),
        ("ITB-066", "Fluid-Applied Flooring — scope removed by the Owner in a value engineering "
                    "exercise (08.04.26). No subcontract will be issued. Those locations became "
                    "sealed concrete under ITB-067."),
    ]
    for i, (a, b) in enumerate(basis, 1):
        ws.cell(i, 1, a).font = BOLD
        ws.cell(i, 1).alignment = WRAP
        ws.cell(i, 2, b).alignment = WRAP
        ws.row_dimensions[i].height = 15 if not b else max(15, 13 * (len(b) // 100 + 1))

    f = OUT / "Buyout Log & Subcontract Release Priority.xlsx"
    wb.save(f)
    print(f"wrote {f.relative_to(ROOT)}")
    return f


if __name__ == "__main__":
    rows = build()
    write_workbook(rows)
    print(f"{len(rows)} packages, ${sum(r['carried_value'] for r in rows):,.0f} carried")
    for r in rows[:8]:
        print(f"  {str(r['priority_rank']):>3}. {r['package_id']}  {r['carried_sub'][:28]:30} "
              f"${r['carried_value']:>12,.0f}  release by {r['latest_release_date']}  "
              f"{r['working_days_late']:>4} d late")
